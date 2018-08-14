import openpyxl as pxl
from openpyxl.styles.borders import Border, Side
import numpy as np

'''range 25 tot 46 wanneer geen info onder prijs, aantal:h, prijs:j, totaal: k
opmaak kolom j afhankelijk van cellen'''
datum = r'10/10/2010'
offertnr = 3078
openlines = 0
listsumlet = ['G','H', 'I', 'J', 'K']

wb = pxl.load_workbook(r'C:\Users\Tom\Documents\Offerte\offerte2.xlsx')
ws = wb.active

def createsheet(filename):

    ws['J16']= datum
    ws['J17'] = offertnr

    ws['A25'] = 'Buitenvloer'
    ws['B26'] = r'- Opmerkingen'
    ws['H25'] = 8
    ws['I25'] = 'x'
    ws['J25'] = 5
    first = ws['H25']
    second = ws['J25']
    ws['K25'] = first.value*second.value
    cell = ws['K25']
    cell.number_format = '€ ##,##'
    cell = ws['J25']
    cell.number_format = '€ ##,##'
    cell = ws['H25']
    cell.number_format = '#,## u'

    #ws['G13'] = ws['A16'].value
    top_border = Border(top = Side(style='thick'))
    bottom_border = Border(bottom=Side(style='thick'))
    #ws['G13'].border = top_border


    #for i in range(items):
     #   additem(i)




    sumarr = [[[],[],[],[],[]],[[],[],[],[],[]],[[],[],[],[],[]]]  # 3x5
    for i in range(5):  # create array met totaalberekening, einde
        for b in range(3):
            if i == 0 or i == 4:
                sumarr[b][i] = ws[listsumlet[i]+str(b+47)].value



    img = pxl.drawing.image.Image(r'C:\Users\Tom\Documents\Offerte\image001.jpg')
    ws.add_image(img, 'A3')

    wb.save(filename)
createsheet(r'C:\Users\Tom\Documents\Offerte\offerte2.xlsx')