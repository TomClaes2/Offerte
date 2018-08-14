import openpyxl as pxl
from openpyxl.styles.borders import Border, Side
import numpy as np


'''items kunnen toegevoegd worden vanaf 25 tot 46,
wanneer lengte van file nadert naar 46, bv 41 of meer: notitie weergeven aan gebruiker en
optie geven om overige lijnen te gebruiken of om nieuwe pagina te starten.'''

pathorigin = r'C:\Users\Tom\Documents\Offerte\offerte2.xlsx'
wb = pxl.load_workbook(pathorigin)
ws = wb.active

'''returnt rij met eerste vrije spatie'''
def lastitem(filename):
    for i in range(22):
        if ws['K'+str(i+25)].value ==None:
            if ws['B'+str(i+25)].value ==None:
                break
    return(i+25)

'''items is lijst met elk item een lijn, eerste lijn begint tegen kantlijn, andere hebben indent'''
def additem(filename, items):
    last = lastitem(filename)
    if last >= 43:
        print('volgende pagina')
    else:
        for i in range(len(items)):
            if i == 0:
                ws['A'+str(last)] = items[0]
            else:
                ws['B'+str(i+last)] = items[i]
    wb.save(filename)

print(additem(r'C:\Users\Tom\Documents\Offerte\offerte2.xlsx',['Binnenvloer', '-Opmerkingen']))
